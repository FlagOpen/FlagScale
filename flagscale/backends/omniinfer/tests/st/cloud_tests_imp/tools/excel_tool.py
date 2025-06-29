# SPDX-License-Identifier: MIT
# Copyright (c) 2025 Huawei Technologies Co., Ltd. All Rights Reserved.

# -*- coding: UTF-8 -*-
# 读取Excel文件
import os
import openpyxl
from openpyxl import load_workbook
import copy

import logging

logger = logging.getLogger(__name__)


def read_excel(path):
    print("read_excel-----------------------------------")
    # 获取工作簿
    book = openpyxl.load_workbook(path)

    # 获取活动行（非空白的）
    sheet = book[os.getenv("case_time")]
    # print(wb.get_sheet_by_name('Sheet1'))

    # 提取数据，格式：[[1, 2, 3], [3, 6, 9], [100, 200, 300]]
    values = []
    case_model = os.getenv("case_type")
    model_architecture = os.getenv("model_architecture")

    try:
        for row in sheet:
            line = []
            if row[0].value is None:  # 如果为空直接就退出，因此表格允许空的用例名字和空行
                continue
            if case_model in row[1].value:
                for cell in row:
                    line.append(cell.value)
            else:
                continue

            values.append(line)
    except Exception as e:
        logger.info(row)
        logger.exception(e)

    return values


def read_excel_new(file_path):
    # 加载工作簿
    workbook = load_workbook(filename=file_path)
    # 获取sheet
    sheet = workbook[os.getenv("case_time")]
    case_model = os.getenv("case_type")
    model_architecture = os.getenv("model_architecture")

    # 查找目标列索引
    column_header = {}
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for index, cell_value in enumerate(row, start=1):
            column_header[cell_value] = index

    lines = []
    i = 0
    for row in sheet:
        if i == 0:  # 不取表头
            i += 1
            continue
        if row[0].value is None:  # 如果为空跳过，因此表格允许空的用例名字和空行
            continue
        if row[column_header.get("用例名称", 1) - 1].value.startswith("#"):
            continue
        case_models = row[column_header.get("模型", 2) - 1].value
        model_architectures = row[column_header.get("形态", 3) - 1].value
        copied_dict = copy.deepcopy(column_header)
        case_models_array = case_models.split(",")

        if case_model in case_models_array:
            if model_architecture in model_architectures.split(",") or "ALL" == model_architectures:
                for key, value in enumerate(column_header):
                    copied_dict[value] = row[key].value
                lines.append(copied_dict)

    return lines


if __name__ == '__main__':
    result = read_excel("1.xlsx")
    print(result)
